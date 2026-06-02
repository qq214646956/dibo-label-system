"""
地博标签打印系统 — 生产服务
Flask 单进程托管：SAP RFC 代理 + MySQL 模板共享 + 前端静态文件
启动: python sap_server.py
"""

import os
import sys
import json
import uuid
from flask import Flask, request, jsonify, send_from_directory
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
from flask_cors import CORS
HAS_SAP = True  # 默认尝试连接
try:
    import mysql.connector
    from mysql.connector import pooling
    HAS_MYSQL = True
except ImportError:
    HAS_MYSQL = False

# =============================================
# 配置
# =============================================

SAP_CONFIG = {
    'ashost': '172.168.10.33',
    'sysnr': '00',
    'client': '100',
    'user': 'IT01',
    'passwd': 'xiaoxiang123',
    'lang': 'ZH'
}

def load_config():
    # 优先读取 exe 同级目录的 config.json
    cfg_file = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__), 'config.json')
    if os.path.exists(cfg_file):
        with open(cfg_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

_cfg = load_config()

MYSQL_CONFIG = {
    'host': _cfg.get('mysql_host', 'localhost'),
    'port': _cfg.get('mysql_port', 3306),
    'user': _cfg.get('mysql_user', 'root'),
    'password': _cfg.get('mysql_password', 'Mysql@2026'),
    'database': _cfg.get('mysql_database', 'label_system'),
    'charset': 'utf8mb4',
    'auth_plugin': _cfg.get('mysql_auth', 'caching_sha2_password'),
    'use_pure': True
}

# 前端静态文件目录（npm run build 的输出）
# PyInstaller 兼容：exe 同级目录的 dist/ 文件夹
if getattr(sys, 'frozen', False):
    DIST_DIR = os.path.join(os.path.dirname(sys.executable), 'dist')
else:
    DIST_DIR = os.path.join(os.path.dirname(__file__), 'dist')

# =============================================
# 初始化
# =============================================

app = Flask(__name__, static_folder=DIST_DIR, static_url_path='')
CORS(app)

# --- 首次启动自动建库建表 ---
def init_database():
    try:
        import mysql.connector
        # 先连接不指定库，创建数据库
        cfg = MYSQL_CONFIG.copy()
        dbname = cfg.pop('database', 'label_system')
        conn = mysql.connector.connect(**cfg)
        cur = conn.cursor()
        cur.execute(f"CREATE DATABASE IF NOT EXISTS `{dbname}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        cur.close()
        conn.close()
        # 再连接指定库，建表
        cfg['database'] = dbname
        conn = mysql.connector.connect(**cfg)
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS label_templates (
            id VARCHAR(36) PRIMARY KEY, name VARCHAR(200) NOT NULL,
            target_entity VARCHAR(50) DEFAULT 'delivery', layout_json LONGTEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4''')
        cur.execute('''CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY, username VARCHAR(50) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL, display_name VARCHAR(100) NOT NULL,
            role ENUM("admin","designer","operator") NOT NULL DEFAULT "operator",
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4''')
        cur.execute('''CREATE TABLE IF NOT EXISTS template_logs (
            id INT AUTO_INCREMENT PRIMARY KEY, template_id VARCHAR(36) NOT NULL,
            template_name VARCHAR(200) NOT NULL, action VARCHAR(20) NOT NULL,
            operator VARCHAR(100) DEFAULT '', created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_template (template_id), INDEX idx_created (created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4''')
        # 建默认管理员（如果用户表为空）
        cur.execute("SELECT COUNT(*) FROM users")
        if cur.fetchone()[0] == 0:
            import hashlib
            pw = hashlib.sha256('admin123'.encode()).hexdigest()
            cur.execute("INSERT INTO users (username, password_hash, display_name, role) VALUES (%s,%s,%s,%s)",
                       ('admin', pw, '管理员', 'admin'))
        conn.commit()
        cur.close()
        conn.close()
        print(" =  数据库初始化完成")
    except Exception as e:
        print(f" =  MySQL 连接失败: {e}")
        print(f" =  请检查: 1)MySQL是否运行 2)config.json密码是否正确")
        print(f" =  当前配置: host={MYSQL_CONFIG['host']} port={MYSQL_CONFIG['port']} user={MYSQL_CONFIG['user']}")

init_database()

# MySQL 连接池（连接失败不阻塞启动）
mysql_pool = None
if HAS_MYSQL:
    try:
        mysql_pool = pooling.MySQLConnectionPool(
            pool_name='label_pool',
            pool_size=5,
            **MYSQL_CONFIG
        )
        print(" =  MySQL 已连接")
    except Exception as e:
        print(f" =  MySQL 未连接，模板功能不可用: {e}")
else:
    print(" =  MySQL 驱动未安装，模板功能不可用")


def get_db():
    if mysql_pool is None:
        raise Exception("MySQL 未连接")
    return mysql_pool.get_connection()


def get_sap_connection():
    if not HAS_SAP:
        return None, "SAP 连接已禁用"
    try:
        # 让 pyrfc 能找到同目录下的 SAP DLL
        import pyrfc
        pyrfc_dir = os.path.dirname(pyrfc.__file__)
        if hasattr(os, 'add_dll_directory'):
            os.add_dll_directory(pyrfc_dir)
        else:
            os.environ['PATH'] = pyrfc_dir + ';' + os.environ.get('PATH', '')
        from pyrfc import Connection
        conn = Connection(**SAP_CONFIG)
        return conn, None
    except ImportError:
        return None, "服务器未安装 SAP RFC 依赖（pyrfc）"
    except Exception as e:
        return None, f"SAP 连接失败: {e}"


def safe_value(val, default=''):
    if val is None:
        return default
    if isinstance(val, str):
        return val.strip()
    return val


# =============================================
# 模板 API
# =============================================

@app.route('/api/templates', methods=['GET'])
def get_templates():
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id, name, target_entity, layout_json FROM label_templates ORDER BY updated_at DESC")
        rows = cur.fetchall()
        templates = []
        for r in rows:
            t = json.loads(r['layout_json'])
            t['id'] = r['id']
            t['name'] = r['name']
            t['targetEntity'] = r['target_entity']
            templates.append(t)
        cur.close()
        db.close()
        return jsonify({'success': True, 'data': templates})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取模板失败: {e}'}), 500


def log_template_action(tid, name, action, operator):
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO template_logs (template_id, template_name, action, operator) VALUES (%s,%s,%s,%s)",
                    (tid, name, action, operator))
        db.commit()
        cur.close()
        db.close()
    except:
        pass


@app.route('/api/templates', methods=['POST'])
def save_template():
    try:
        data = request.json
        tid = data['id']
        name = data.get('name', '')
        target_entity = data.get('targetEntity', 'delivery')
        operator = request.args.get('operator', '')
        layout_json = json.dumps(data, ensure_ascii=False)

        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT id FROM label_templates WHERE id=%s", (tid,))
        existed = cur.fetchone() is not None
        cur.execute(
            "INSERT INTO label_templates (id, name, target_entity, layout_json) "
            "VALUES (%s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE name=%s, target_entity=%s, layout_json=%s",
            (tid, name, target_entity, layout_json, name, target_entity, layout_json)
        )
        db.commit()
        cur.close()
        db.close()
        log_template_action(tid, name, 'UPDATE' if existed else 'CREATE', operator)
        return jsonify({'success': True, 'id': tid})
    except Exception as e:
        return jsonify({'success': False, 'message': f'保存模板失败: {e}'}), 500


@app.route('/api/templates/<tid>', methods=['DELETE'])
def delete_template(tid):
    try:
        operator = request.args.get('operator', '')
        # Get name before deleting for log
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT name FROM label_templates WHERE id=%s", (tid,))
        row = cur.fetchone()
        tname = row[0] if row else tid
        cur.execute("DELETE FROM label_templates WHERE id = %s", (tid,))
        db.commit()
        cur.close()
        db.close()
        log_template_action(tid, tname, 'DELETE', operator)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': f'删除模板失败: {e}'}), 500


@app.route('/api/template-logs', methods=['GET'])
def get_template_logs():
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM template_logs ORDER BY created_at DESC LIMIT 200")
        logs = cur.fetchall()
        for l in logs:
            l['created_at'] = l['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        cur.close()
        db.close()
        return jsonify({'success': True, 'data': logs})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =============================================
# 用户认证 API
# =============================================

import hashlib

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username or not password:
        return jsonify({'success': False, 'message': '请输入用户名和密码'}), 400

    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT id, username, display_name, role FROM users WHERE username=%s AND password_hash=%s",
                (username, pw_hash))
    user = cur.fetchone()
    cur.close()
    db.close()

    if user:
        return jsonify({'success': True, 'user': user})
    return jsonify({'success': False, 'message': '用户名或密码错误'}), 401


@app.route('/api/users', methods=['GET'])
def get_users():
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT id, username, display_name, role, created_at FROM users ORDER BY id")
    users = cur.fetchall()
    cur.close()
    db.close()
    return jsonify({'success': True, 'data': users})


@app.route('/api/users', methods=['POST'])
def create_user():
    import hashlib
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    display_name = data.get('display_name', '').strip()
    role = data.get('role', 'operator').strip()

    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400

    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("INSERT INTO users (username, password_hash, display_name, role) VALUES (%s,%s,%s,%s)",
                    (username, pw_hash, display_name, role))
        db.commit()
        cur.close()
        db.close()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except mysql.connector.errors.IntegrityError:
        return jsonify({'success': False, 'message': '用户名已存在'}), 400


@app.route('/api/users/<int:uid>', methods=['DELETE'])
def delete_user(uid):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT role FROM users WHERE id=%s", (uid,))
    row = cur.fetchone()
    if row and row[0] == 'admin':
        # Check if this is the last admin
        cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
        if cur.fetchone()[0] <= 1:
            cur.close()
            db.close()
            return jsonify({'success': False, 'message': '不能删除最后一个管理员'}), 400
    cur.execute("DELETE FROM users WHERE id=%s", (uid,))
    db.commit()
    cur.close()
    db.close()
    return jsonify({'success': True})


# =============================================
# SAP RFC API
# =============================================

@app.route('/api/delivery-details', methods=['GET'])
def delivery_details():
    iv_wbstk = request.args.get('iv_wbstk', '').strip()
    iv_cust_name = request.args.get('iv_cust_name', '').strip()
    iv_wadat_from = request.args.get('iv_wadat_from', '').strip()
    iv_wadat_to = request.args.get('iv_wadat_to', '').strip()
    iv_erdat_from = request.args.get('iv_erdat_from', '').strip()
    iv_erdat_to = request.args.get('iv_erdat_to', '').strip()

    conn, error = get_sap_connection()
    if conn is None:
        return jsonify({'success': False, 'message': error}), 500

    try:
        result = conn.call(
            'ZFM_ZSDELIVERY_DETAILS',
            IV_WBSTK=iv_wbstk,
            IV_CUST_NAME=iv_cust_name,
            IV_WADAT_FROM=iv_wadat_from,
            IV_WADAT_TO=iv_wadat_to,
            IV_ERDAT_FROM=iv_erdat_from,
            IV_ERDAT_TO=iv_erdat_to,
            ET_OUTPUT=[]
        )

        raw_data = result.get('ET_OUTPUT', [])
        records = []
        for item in raw_data:
            records.append({
                'VBELN': safe_value(item.get('VBELN')),
                'POSNR': safe_value(item.get('POSNR')),
                'VGBEL': safe_value(item.get('VGBEL')),
                'VGPOS': safe_value(item.get('VGPOS')),
                'MATNR': safe_value(item.get('MATNR')),
                'KDMAT': safe_value(item.get('KDMAT')),
                'ARKTX': safe_value(item.get('ARKTX')),
                'LFIMG': safe_value(item.get('LFIMG')),
                'MEINS': safe_value(item.get('MEINS')),
                'CHARG': safe_value(item.get('CHARG')),
                'KUNNR': safe_value(item.get('KUNNR')),
                'ERDAT': safe_value(item.get('ERDAT')),
                'WADAT_IST': safe_value(item.get('WADAT_IST')),
                'GROES': safe_value(item.get('GROES')),
                'MAKTX': safe_value(item.get('MAKTX')),
                'PM': safe_value(item.get('PM')),
                'ZBZCD': safe_value(item.get('ZBZCD')),
                'ZBZKD': safe_value(item.get('ZBZKD')),
                'ZBZHD': safe_value(item.get('ZBZHD')),
                'ZBZBZ': safe_value(item.get('ZBZBZ')),
                'ZBZZL': safe_value(item.get('ZBZZL')),
                'ZHYZL': safe_value(item.get('ZHYZL')),
                'ZFZDW': safe_value(item.get('ZFZDW')),
                'ZBZCD2': safe_value(item.get('ZBZCD2')),
                'ZBZKD2': safe_value(item.get('ZBZKD2')),
                'ZFZSL': safe_value(item.get('ZFZSL')),
                'LFIMG_HY': safe_value(item.get('LFIMG_HY')),
                'SORTL': safe_value(item.get('SORTL')),
                'NAME1': safe_value(item.get('NAME1')),
                'USERNAME': safe_value(item.get('USERNAME')),
            })

        return jsonify({
            'success': True,
            'count': len(records),
            'data': records
        })

    except (ABAPApplicationError, ABAPRuntimeError) as e:
        return jsonify({'success': False, 'message': f'SAP RFC 执行错误: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'调用失败: {e}'}), 500
    finally:
        conn.close()


# =============================================
# 前端静态文件
# =============================================

@app.route('/')
def index():
    return send_from_directory(DIST_DIR, 'index.html')


@app.route('/<path:path>')
def static_files(path):
    file_path = os.path.join(DIST_DIR, path)
    if os.path.isfile(file_path):
        return send_from_directory(DIST_DIR, path)
    return send_from_directory(DIST_DIR, 'index.html')


# =============================================
# 启动
# =============================================

# =============================================
# Excel 导入 API
# =============================================

@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    if not HAS_OPENPYXL:
        return jsonify({'success': False, 'message': '服务器未安装 openpyxl，请执行 pip install openpyxl'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上传文件'}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传 .xlsx 文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        merged_ranges = list(ws.merged_cells.ranges)
        elements = []
        processed_cells = set()

        for merged in merged_ranges:
            min_row, min_col = merged.min_row, merged.min_col
            max_row, max_col = merged.max_row, merged.max_col
            cell = ws.cell(min_row, min_col)
            content = str(cell.value).strip() if cell.value is not None else ''
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    processed_cells.add((r, c))
            style = _extract_merged_style(ws, min_row, min_col, max_row, max_col)
            if not content and not style:
                continue
            x_mm, y_mm = _excel_to_mm_pos(ws, min_row, min_col)
            w_mm = _col_range_width_mm(ws, min_col, max_col)
            h_mm = _row_range_height_mm(ws, min_row, max_row)
            elem = {
                'id': str(uuid.uuid4())[:8], 'type': 'text',
                'x': round(x_mm, 1), 'y': round(y_mm, 1),
                'w': round(w_mm, 1), 'h': round(h_mm, 1),
                'content': content, 'style': style if style else {}
            }
            elements.append(elem)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if (cell.row, cell.column) in processed_cells:
                    continue
                content = str(cell.value).strip() if cell.value is not None else ''
                style = _extract_cell_style(cell)
                if not content and not style:
                    continue
                x_mm, y_mm = _excel_to_mm_pos(ws, cell.row, cell.column)
                w_mm = _col_width_mm(ws, cell.column)
                h_mm = _row_height_mm(ws, cell.row)
                elem = {
                    'id': str(uuid.uuid4())[:8], 'type': 'text',
                    'x': round(x_mm, 1), 'y': round(y_mm, 1),
                    'w': round(w_mm, 1), 'h': round(h_mm, 1),
                    'content': content, 'style': style if style else {}
                }
                elements.append(elem)
                processed_cells.add((cell.row, cell.column))

        if not elements:
            return jsonify({'success': False, 'message': 'Excel 中没有找到任何内容'}), 400

        max_right = max(e['x'] + e['w'] for e in elements)
        max_bottom = max(e['y'] + e['h'] for e in elements)
        template_name = file.filename.rsplit('.', 1)[0]

        layout = {
            'id': str(uuid.uuid4())[:8], 'name': template_name,
            'width': round(max_right, 1), 'height': round(max_bottom, 1),
            'unit': 'mm', 'targetEntity': 'delivery',
            'backgroundColor': '#ffffff', 'elements': elements
        }
        wb.close()
        return jsonify({'success': True, 'data': layout})
    except Exception as e:
        return jsonify({'success': False, 'message': f'解析 Excel 失败: {e}'}), 500


def _col_width_mm(ws, col):
    letter = get_column_letter(col)
    cd = ws.column_dimensions.get(letter)
    char_units = cd.width if (cd and cd.width) else 8.43
    px = char_units * 12 if char_units <= 1 else char_units * 7 + 5
    return px / 96 * 25.4

def _row_height_mm(ws, row):
    rd = ws.row_dimensions.get(row)
    pt = rd.height if (rd and rd.height) else 15
    return pt * 0.3528

def _col_range_width_mm(ws, min_col, max_col):
    return sum(_col_width_mm(ws, c) for c in range(min_col, max_col + 1))

def _row_range_height_mm(ws, min_row, max_row):
    return sum(_row_height_mm(ws, r) for r in range(min_row, max_row + 1))

def _excel_to_mm_pos(ws, row, col):
    x_mm = sum(_col_width_mm(ws, c) for c in range(1, col))
    y_mm = sum(_row_height_mm(ws, r) for r in range(1, row))
    return x_mm, y_mm

def _cell_has_border(cell):
    """检查单个单元格是否有边框"""
    if not cell.border:
        return False
    sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
    return any(s and s.style is not None for s in sides)


def _extract_cell_style(cell):
    """提取单单元格样式，无边框返回 None"""
    style = {}
    has_any = False

    if cell.font:
        if cell.font.size:
            style['fontSize'] = int(cell.font.size)
            has_any = True
        if cell.font.bold:
            style['fontWeight'] = 'bold'
            has_any = True
        if cell.font.color and cell.font.color.rgb:
            try:
                rgb = cell.font.color.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    style['color'] = f'#{rgb}'
                    has_any = True
            except Exception:
                pass
    if cell.alignment:
        h, v = cell.alignment.horizontal, cell.alignment.vertical
        if h in ('left', 'center', 'right'):
            style['textAlign'] = h
            has_any = True
        if v in ('top', 'middle', 'bottom', 'center'):
            style['verticalAlign'] = v if v != 'center' else 'middle'
            has_any = True
    if _cell_has_border(cell):
        style['borderWidth'] = 1
        style['borderColor'] = '#000000'
        style['borderStyle'] = 'solid'
        has_any = True

    if not has_any:
        return None
    style.setdefault('fontSize', 10)
    return style


def _extract_merged_style(ws, min_row, min_col, max_row, max_col):
    """提取合并单元格样式：从四边收集边框"""
    cell = ws.cell(min_row, min_col)
    style = _extract_cell_style(cell)
    if style is None:
        style = {}

    # 收集合并区域外围边框（四条边各自的任意子单元格有边框即算有边框）
    has_border = False

    # 上边
    for c in range(min_col, max_col + 1):
        if _cell_has_border(ws.cell(min_row, c)):
            has_border = True
            break
    # 下边
    if not has_border:
        for c in range(min_col, max_col + 1):
            if _cell_has_border(ws.cell(max_row, c)):
                has_border = True
                break
    # 左边
    if not has_border:
        for r in range(min_row, max_row + 1):
            if _cell_has_border(ws.cell(r, min_col)):
                has_border = True
                break
    # 右边
    if not has_border:
        for r in range(min_row, max_row + 1):
            if _cell_has_border(ws.cell(r, max_col)):
                has_border = True
                break

    if has_border:
        style['borderWidth'] = 1
        style['borderColor'] = '#000000'
        style['borderStyle'] = 'solid'

    style.setdefault('fontSize', 10)
    if len(style) <= 2:  # only fontSize (+ maybe one more default)
        return None
    return style


if __name__ == '__main__':
    from waitress import serve
    print(f" ========== 地博标签打印系统 ==========")
    print(f" =  地址: http://localhost:5000")
    print(f" =  API: /api/delivery-details")
    print(f" =  模板: /api/templates (MySQL)")
    print(f" =  前端: {DIST_DIR}")
    print(f" =========================================")
    serve(app, host='0.0.0.0', port=5000)
